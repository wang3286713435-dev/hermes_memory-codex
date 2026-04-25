from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.parsing.base import ParsedBlock, ParsedDocument


class XlsxParser:
    def __init__(self, rows_per_block: int = 20) -> None:
        self.rows_per_block = rows_per_block

    def parse(self, source_path: Path) -> ParsedDocument:
        workbook = load_workbook(source_path, data_only=False, read_only=False)
        blocks: list[ParsedBlock] = []

        for sheet in workbook.worksheets:
            rows = self._non_empty_rows(sheet)
            if not rows:
                continue
            headers = self._headers(rows)
            for offset in range(0, len(rows), self.rows_per_block):
                group = rows[offset : offset + self.rows_per_block]
                cell_range = self._cell_range(group)
                text = self._format_group(sheet.title, group, headers, cell_range)
                row_start = group[0]["row"]
                row_end = group[-1]["row"]
                col_start = min(cell["column"] for row in group for cell in row["cells"])
                col_end = max(cell["column"] for row in group for cell in row["cells"])
                metadata = {
                    "parser": "xlsx",
                    "sheet_name": sheet.title,
                    "cell_range": cell_range,
                    "row_start": row_start,
                    "row_end": row_end,
                    "column_start": get_column_letter(col_start),
                    "column_end": get_column_letter(col_end),
                    "row_count": len(group),
                    "column_count": col_end - col_start + 1,
                    "headers": headers,
                    "formula_refs": self._formula_refs(group),
                    "numeric_values": self._numeric_values(group),
                    "units": self._units(group),
                    "citation_label": f"{source_path.name} / {sheet.title} / {cell_range}",
                    "chunk_boundary": True,
                }
                blocks.append(
                    ParsedBlock(
                        text=text,
                        block_type="spreadsheet_range",
                        title_path=[f"Sheet: {sheet.title}", f"Range: {cell_range}"],
                        metadata=metadata,
                    )
                )

        return ParsedDocument(
            source_path=source_path,
            title=source_path.name,
            blocks=blocks,
            metadata={"parser": "xlsx", "sheet_count": len(workbook.worksheets)},
        )

    def _non_empty_rows(self, sheet: Any) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for row in sheet.iter_rows():
            cells = []
            for cell in row:
                value = cell.value
                if value is None or str(value).strip() == "":
                    continue
                cells.append(
                    {
                        "coordinate": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "value": value,
                    }
                )
            if cells:
                rows.append({"row": cells[0]["row"], "cells": cells})
        return rows

    def _headers(self, rows: list[dict[str, Any]]) -> list[str]:
        first = rows[0]["cells"] if rows else []
        headers = [self._clean_value(cell["value"]) for cell in first]
        return [header for header in headers if header]

    def _cell_range(self, rows: list[dict[str, Any]]) -> str:
        min_row = min(row["row"] for row in rows)
        max_row = max(row["row"] for row in rows)
        min_col = min(cell["column"] for row in rows for cell in row["cells"])
        max_col = max(cell["column"] for row in rows for cell in row["cells"])
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    def _format_group(
        self,
        sheet_name: str,
        rows: list[dict[str, Any]],
        headers: list[str],
        cell_range: str,
    ) -> str:
        lines = [f"Sheet: {sheet_name}", f"Cell range: {cell_range}"]
        if headers:
            lines.append("Headers: " + " | ".join(headers))
        for row in rows:
            parts = []
            for cell in row["cells"]:
                value = self._clean_value(cell["value"])
                if str(cell["value"]).startswith("="):
                    value = f"formula {value}"
                parts.append(f"{cell['coordinate']}={value}")
            lines.append(f"Row {row['row']}: " + " | ".join(parts))
        return "\n".join(lines)

    def _formula_refs(self, rows: list[dict[str, Any]]) -> list[dict[str, str]]:
        formulas = []
        for row in rows:
            for cell in row["cells"]:
                value = str(cell["value"])
                if value.startswith("="):
                    formulas.append({"cell": cell["coordinate"], "formula": value})
        return formulas

    def _numeric_values(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        values = []
        for row in rows:
            for cell in row["cells"]:
                value = cell["value"]
                if isinstance(value, int | float):
                    values.append({"cell": cell["coordinate"], "value": str(value)})
        return values

    def _units(self, rows: list[dict[str, Any]]) -> list[str]:
        known_units = ("元", "万元", "%", "天", "日", "m2", "㎡", "台", "套", "项")
        found: list[str] = []
        for row in rows:
            for cell in row["cells"]:
                value = self._clean_value(cell["value"])
                for unit in known_units:
                    if unit in value and unit not in found:
                        found.append(unit)
        return found

    def _clean_value(self, value: Any) -> str:
        return str(value).replace("\n", " ").strip()
